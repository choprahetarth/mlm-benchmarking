#%%
from datasets import load_dataset
dataset = load_dataset("hotpotqa/hotpot_qa", "distractor", split='train', trust_remote_code=True , streaming=True)
print(next(iter(dataset)))

# - id: a unique identifier for the question
# - question: the question being asked
# - answer: the answer to the question
# - type: the type of question (e.g. "bridge", "comparison", etc.)
# - level: the difficulty level of the question
# - supporting_facts: a list of sentences that support the answer
# - context: a list of paragraphs related to the question

# {'id': '5a7a06935542990198eaf050',
# 'question': "Which magazine was started first Arthur's Magazine or First for Women?",
# 'answer': "Arthur's Magazine",
# 'type': 'comparison',
# 'level': 'medium',
# 'supporting_facts': {'title': ["Arthur's Magazine", 'First for Women'], 'sent_id': [0, 0]},
# 'context': {'title': ['Radio City (Indian radio station)', 'History of Albanian football', 'Echosmith', "Women's colleges in the Southern United States", 'First Arthur County Courthouse and Jail', "Arthur's Magazine", '2014–15 Ukrainian Hockey Championship', 'First for Women', 'Freeway Complex Fire', 'William Rast'], 'sentences': [["Radio City is India's first private FM radio station and was started on 3 July 2001.", ' It broadcasts on 91.1 (earlier 91.0 in most cities) megahertz from Mumbai (where it was started in 2004), Bengaluru (started first in 2001), Lucknow and New Delhi (since 2003).', ' It plays Hindi, English and regional songs.', ' It was launched in Hyderabad in March 2006, in Chennai on 7 July 2006 and in Visakhapatnam October 2007.', ' Radio City recently forayed into New Media in May 2008 with the launch of a music portal - PlanetRadiocity.com that offers music related news, videos, songs, and other music-related features.', ' The Radio station currently plays a mix of Hindi and Regional music.', ' Abraham Thomas is the CEO of the company.'], ['Football in Albania existed before the Albanian Football Federation (FSHF) was created.', " This was evidenced by the team's registration at the Balkan Cup tournament during 1929-1931, which started in 1929 (although Albania eventually had pressure from the teams because of competition, competition started first and was strong enough in the duels) .", ' Albanian National Team was founded on June 6, 1930, but Albania had to wait 16 years to play its first international match and then defeated Yugoslavia in 1946.', ' In 1932, Albania joined FIFA (during the 12–16 June convention ) And in 1954 she was one of the founding members of UEFA.'], ['Echosmith is an American, Corporate indie pop band formed in February 2009 in Chino, California.', ' Originally formed as a quartet of siblings, the band currently consists of Sydney, Noah and Graham Sierota, following the departure of eldest sibling Jamie in late 2016.', ' Echosmith started first as "Ready Set Go!"', ' until they signed to Warner Bros.', ' Records in May 2012.', ' They are best known for their hit song "Cool Kids", which reached number 13 on the "Billboard" Hot 100 and was certified double platinum by the RIAA with over 1,200,000 sales in the United States and also double platinum by ARIA in Australia.', ' The song was Warner Bros.', " Records' fifth-biggest-selling-digital song of 2014, with 1.3 million downloads sold.", ' The band\'s debut album, "Talking Dreams", was released on October 8, 2013.'], ["Women's colleges in the Southern United States refers to undergraduate, bachelor's degree–granting institutions, often liberal arts colleges, whose student populations consist exclusively or almost exclusively of women, located in the Southern United States.", " Many started first as girls' seminaries or academies.", ' Salem College is the oldest female educational institution in the South and Wesleyan College is the first that was established specifically as a college for women.', ' Some schools, such as Mary Baldwin University and Salem College, offer coeducational courses at the graduate level.'], ['The First Arthur County Courthouse and Jail, was perhaps the smallest court house in the United States, and serves now as a museum.'], ["Arthur's Magazine (1844–1846) was an American literary periodical published in Philadelphia in the 19th century.", ' Edited by T.S. Arthur, it featured work by Edgar A. Poe, J.H. Ingraham, Sarah Josepha Hale, Thomas G. Spear, and others.', ' In May 1846 it was merged into "Godey\'s Lady\'s Book".'], ['The 2014–15 Ukrainian Hockey Championship was the 23rd season of the Ukrainian Hockey Championship.', ' Only four teams participated in the league this season, because of the instability in Ukraine and that most of the clubs had economical issues.', ' Generals Kiev was the only team that participated in the league the previous season, and the season started first after the year-end of 2014.', ' The regular season included just 12 rounds, where all the teams went to the semifinals.', ' In the final, ATEK Kiev defeated the regular season winner HK Kremenchuk.'], ["First for Women is a woman's magazine published by Bauer Media Group in the USA.", ' The magazine was started in 1989.', ' It is based in Englewood Cliffs, New Jersey.', ' In 2011 the circulation of the magazine was 1,310,696 copies.'], ['The Freeway Complex Fire was a 2008 wildfire in the Santa Ana Canyon area of Orange County, California.', ' The fire started as two separate fires on November 15, 2008.', ' The "Freeway Fire" started first shortly after 9am with the "Landfill Fire" igniting approximately 2 hours later.', ' These two separate fires merged a day later and ultimately destroyed 314 residences in Anaheim Hills and Yorba Linda.'], ['William Rast is an American clothing line founded by Justin Timberlake and Trace Ayala.', ' It is most known for their premium jeans.', ' On October 17, 2006, Justin Timberlake and Trace Ayala put on their first fashion show to launch their new William Rast clothing line.', ' The label also produces other clothing items such as jackets and tops.', ' The company started first as a denim line, later evolving into a men’s and women’s clothing line.']]}}

#%%
from datasets import load_dataset
from PIL import Image
import io


dataset = load_dataset("HuggingFaceM4/A-OKVQA",split='train',  trust_remote_code=True , streaming=True)
sample = next(iter(dataset))
print(sample)
image = sample['image']
image.show()




#'image', 'question_id', 'question', 'choices', 'correct_choice_idx', 'direct_answers', 'difficult_direct_answer', 'rationales'

# - image: image data
# - question_id: unique identifier for each question
# - question: text of the question being asked
# - choices: list of possible answers
# - correct_choice_idx: index of the correct answer in the choices list

# - direct_answers: list of direct answers to the question

# - difficult_direct_answer: boolean indicating whether the direct answer is difficult to determine
# - rationales: list of rationales or explanations for the answer

# {'image': <PIL.JpegImagePlugin.JpegImageFile image mode=RGB size=640x480 at 0x17FEF711070>,
# 'question_id': '22MexNkBPpdZGX6sxbxVBH',
# 'question': 'What is the man by the bags awaiting?',
# 'choices': ['skateboarder', 'train', 'delivery', 'cab'],
# 'correct_choice_idx': 3,
# 'direct_answers': "['ride', 'ride', 'bus', 'taxi', 'travelling', 'traffic', 'taxi', 'cab', 'cab', 'his ride']",
# 'difficult_direct_answer': False,
# 'rationales': ['A train would not be on the street, he would not have luggage waiting for a delivery, and the skateboarder is there and not paying attention to him so a cab is the only possible answer.',
#                 'He has bags as if he is going someone, and he is on a road waiting for vehicle that can only be moved on the road and is big enough to hold the bags.', 'He looks to be waiting for a paid ride to pick him up.']}




#%%


from datasets import load_dataset
import json
import numpy as np
import base64

dataset = load_dataset("HuggingFaceM4/A-OKVQA", split='validation', trust_remote_code=True, streaming=True)

def transform_to_hotpotqa(example):
    """
    The format of output should be like this
    
    - id: a unique identifier for the question
    - question: the question being asked
    - answer: the answer to the question
    - type: the type of question (e.g. "bridge", "comparison", etc.) (still have not figured this out)
    - level: the difficulty level of the question
    - supporting_facts: a list of sentences that support the answer
    - context: a list of paragraphs related to the question
    
    The format of input should be like this 
    
    - image: image data
    - question_id: unique identifier for each question
    - question: text of the question being asked
    - choices: list of possible answers
    - correct_choice_idx: index of the correct answer in the choices list
    - direct_answers: list of direct answers to the question
    - difficult_direct_answer: boolean indicating whether the direct answer is difficult to determine
    - rationales: list of rationales or explanations for the answer
    """
    # Extract relevant information from the A-OKVQA example
    question_id = example['question_id']
    question = example['question']
    choices = example['choices']
    answer = example['choices'][example['correct_choice_idx']]
    rationales = example['rationales']

    context = np.array(example['image']).tolist()
    # context = base64.b64encode(example['image']).decode('utf-8')
    supporting_facts = {
        'Thoughts': rationales,
        'correct_choice_idx': example['correct_choice_idx']
    }
    hotpotqa_example = {
        'id': question_id,
        'question': question + ', Your Choices are - ' + str(choices),
        'answer': answer,
        'type': '',  # This might need to be adjusted based on the actual question type
        'level': 'medium',  # This might need to be adjusted based on the actual difficulty level
        'direct_answers': example['direct_answers'],
        'supporting_facts': supporting_facts,
        'context': context # this needs to be explained by an LLM
    }

    return hotpotqa_example

# Apply the transformation to the dataset and remove the 'image' column
hotpotqa_dataset = dataset.map(transform_to_hotpotqa, remove_columns=['image','question_id', 'rationales', 'difficult_direct_answer', 'correct_choice_idx'])

# Print the first example of the transformed dataset
import json
print(json.dumps(next(iter(hotpotqa_dataset)), indent=1))

#%%
from datasets import load_dataset
import json
import base64
from datasets import get_dataset_config_names
import ast
from io import BytesIO

# List all available configurations for the dataset
configs = get_dataset_config_names('MMMU/MMMU')
print("Available configs:", configs)

def transform_to_hotpotqa_with_multiple_images(example):
    # Extract relevant information
    question_id = example['id']

    # Initialize question_text and question_image
    question_text = ''
    question_image = ''

    # Get the question
    question = example['question']
    if isinstance(question, str):
        question_text = question
    else:
        # If question is an image, encode it to base64
        buffered = BytesIO()
        question.save(buffered, format='PNG')
        question_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        question_image = question_base64

    # Parse the options from string to a list
    options_field = example['options']
    options = []
    if isinstance(options_field, str):
        options = ast.literal_eval(options_field)
    elif isinstance(options_field, list):
        options = options_field
    else:
        # If options_field is an image or other type, handle accordingly
        buffered = BytesIO()
        options_field.save(buffered, format='PNG')
        options_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        options = [options_base64]

    # Map answer letters to indices dynamically
    answer_letter = example['answer'].strip().upper()
    import string
    option_letters = list(string.ascii_uppercase)
    letter_to_index = {letter: idx for idx, letter in enumerate(option_letters)}

    correct_choice_idx = letter_to_index.get(answer_letter, None)
    if correct_choice_idx is None or correct_choice_idx >= len(options):
        correct_choice_idx = None  # Handle invalid or out-of-range indices

    # Get the explanation or rationale
    rationales_field = example.get('explanation', '')
    rationales = ''
    if isinstance(rationales_field, str):
        rationales = rationales_field
    elif rationales_field is not None:
        # If rationales_field is an image, encode it
        buffered = BytesIO()
        rationales_field.save(buffered, format='PNG')
        rationales_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        rationales = rationales_base64

    # Collect all non-None images into a list and encode them as base64 strings
    images = []
    for i in range(1, 8):
        img_key = f'image_{i}'
        img = example.get(img_key, None)
        if img is not None:
            buffered = BytesIO()
            img.save(buffered, format='PNG')
            img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
            images.append(img_base64)

    # Construct the HotpotQA-formatted example
    hotpotqa_example = {
        'id': question_id,
        'question': question_text,
        'question_image': question_image,
        'choices': options,
        'answer': options[correct_choice_idx] if correct_choice_idx is not None else None,
        'type': example.get('question_type', ''),
        'level': example.get('topic_difficulty', 'medium'),
        'explanation': rationales,
        'images': images  # List of base64-encoded image strings
    }

    return hotpotqa_example


# for i in configs:
#     dataset = load_dataset("MMMU/MMMU", i, split='dev', trust_remote_code=True, streaming=True)
#     # example = next(iter(dataset))
#     # transformed_example = transform_to_hotpotqa_with_multiple_images(example)
#     # Apply the transformation to the dataset and remove the 'image' column
#     transformed_dataset = dataset.map(transform_to_hotpotqa_with_multiple_images, remove_columns=['question_image','id', 'supporting_facts'])
#     print(json.dumps(next(iter(transformed_dataset)), indent=1))  
#     break  # Remove this break if you want to process all configs

fields_to_remove = ['question_image', 'id', 'supporting_facts']

all_transformed_examples = []

# for i in configs:
dataset = load_dataset("MMMU/MMMU", "Basic_Medical_Science", split='validation', trust_remote_code=True, streaming=True)
example_count = 0
for example in dataset:
    transformed_example = transform_to_hotpotqa_with_multiple_images(example)
    # Remove specified fields
    for field in fields_to_remove:
        transformed_example.pop(field, None)
    # Add the config name as a field in the final JSON
    transformed_example['config_name'] = "Basic Medical Science"
    # Append the transformed example to the list
    all_transformed_examples.append(transformed_example)
    example_count += 1
        # if example_count >= 5:
            # break  # Take only two samples from each config
    # Continue to the next config without breaking

# Save the collected transformed examples to a JSON file
with open('mmmu_repurposed_val_basic_medical_science.json', 'w') as f:
    json.dump(all_transformed_examples, f, indent=2)

# print(f"Saved {len(all_transformed_examples)} examples to 'transformed_data.json'")

#%%
import pandas as pd
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# Create a DataFrame from the logs
df = pd.DataFrame(all_transformed_examples)
# Convert the 'choices' column from list to string
df['choices'] = df['choices'].apply(lambda x: ', '.join(map(str, x)))


# Create an Excel workbook and remove the default sheet
wb = Workbook()
wb.remove(wb.active)

# For each unique config_name, create a new sheet and write data
for config_name in df['config_name'].unique():
    # Create a new sheet with the config_name
    ws = wb.create_sheet(title=config_name)
    
    # Filter the dataframe for the current config_name
    df_config = df[df['config_name'] == config_name].reset_index(drop=True)
    
    # Remove the 'config_name' column
    df_config = df_config.drop(columns=['config_name'])
    
    # Write the header
    for col_idx, header in enumerate(df_config.columns):
        ws.cell(row=1, column=col_idx + 1, value=header)
    
    # Write the data
    for row_idx, row_data in df_config.iterrows():
        for col_idx, value in enumerate(row_data):
            column_name = df_config.columns[col_idx]
            if column_name == 'images':  # Check if the column is 'images'
                # Check if image_data is a list
                if isinstance(value, list):
                    # Decode each image in the list and save them in the next columns
                    for i, img_data in enumerate(value):
                        image_stream = BytesIO(base64.b64decode(img_data))
                        img = Image(image_stream)
                        aspect_ratio = img.height / img.width
                        img.width = 250
                        img.height = int(250 * aspect_ratio)
                        # Adjust the column index for multiple images
                        img_col = col_idx + 1 + i
                        ws.add_image(img, ws.cell(row=row_idx + 2, column=img_col).coordinate)
                else:
                    # Decode the base64 string and save it as an image
                    image_data = base64.b64decode(value)
                    image_stream = BytesIO(image_data)
                    img = Image(image_stream)
                    aspect_ratio = img.height / img.width
                    img.width = 250
                    img.height = int(250 * aspect_ratio)
                    ws.add_image(img, ws.cell(row=row_idx + 2, column=col_idx + 1).coordinate)
            else:
                if not isinstance(value, list):  # Ensure the value is not a list
                    ws.cell(row=row_idx + 2, column=col_idx + 1, value=str(value))

# Save the workbook
wb.save('output.xlsx')




#%%
from datasets import load_dataset

dataset = load_dataset("HuggingFaceM4/OK-VQA",split='train',  trust_remote_code=True , streaming=True)
print(next(iter(dataset)))

# {'image': <PIL.JpegImagePlugin.JpegImageFile image mode=RGB size=640x479 at 0x1DEFECE8D90>,
# 'question_type': 'four',
# 'confidence': 3,
# 'answers': [{'answer_id': 1, 'raw_answer': 'pony tail', 'answer_confidence': 'yes', 'answer': 'pony tail'},
#             {'answer_id': 2, 'raw_answer': 'pony tail', 'answer_confidence': 'yes', 'answer': 'pony tail'},
#             {'answer_id': 3, 'raw_answer': 'pony tail', 'answer_confidence': 'yes', 'answer': 'pony tail'},
#             {'answer_id': 4, 'raw_answer': 'pony tail', 'answer_confidence': 'yes', 'answer': 'pony tail'},
#             {'answer_id': 5, 'raw_answer': 'pony tail', 'answer_confidence': 'yes', 'answer': 'pony tail'},
#             {'answer_id': 6, 'raw_answer': 'pony tail', 'answer_confidence': 'yes', 'answer': 'pony tail'},
#             {'answer_id': 7, 'raw_answer': 'braid', 'answer_confidence': 'yes', 'answer': 'braid'},
#             {'answer_id': 8, 'raw_answer': 'braid', 'answer_confidence': 'yes', 'answer': 'braid'},
#             {'answer_id': 9, 'raw_answer': 'ponytail', 'answer_confidence': 'yes', 'answer': 'ponytail'},
#             {'answer_id': 10, 'raw_answer': 'ponytail', 'answer_confidence': 'yes', 'answer': 'ponytail'}],
# 'image_id': 51606,
# 'answer_type': 'other',
# 'question_id': 516065,
# 'question': 'What is the hairstyle of the blond called?'}

# %%
from datasets import load_dataset

dataset = load_dataset("google/spiqa",split='test',  trust_remote_code=True , streaming=True)
print(next(iter(dataset)))

# {'answer': 'The Bonaparte school focuses on outdoor physical activities, maneuvers, and strategies, with a specialization in horse riding, lances, and swords. They aim to develop students into good leaders. The Voltaire school, on the other hand, encourages independent thinking and focuses on indoor activities. They aim to instill good moral values and develop students into philosophical thinkers.',
#  'rationale': 'The figure presents a question and answer format, where the answer explicitly describes the contrasting educational approaches of the two schools.',
#  'paper_id': '1611.04684v1',
#  'reference_figure': '1611.04684v1-Table1-1.png',
#  'all_figures': {'1603.00286v5-Figure1-1.png': None, '1603.00286v5-Figure2-1.png': None, '1603.00286v5-Figure3-1.png': None, '1603.00286v5-Figure4-1.png': None, '1603.00286v5-Figure5-1.png': None, '1603.00286v5-Table1-1.png': None, '1603.03833v4-Figure1-1.png': None, '1603.03833v4-Figure2-1.png': None, '1603.03833v4-Figure3-1.png': None, '1603.03833v4-Figure4-1.png': None, '1603.03833v4-Figure5-1.png': None, '1603.03833v4-Table1-1.png': None, '1605.07496v3-Figure1-1.png': None, '1605.07496v3-Figure10-1.png': None, '1605.07496v3-Figure11-1.png': None, '1605.07496v3-Figure2-1.png': None, '1605.07496v3-Figure3-1.png': None, '1605.07496v3-Figure4-1.png': None, '1605.07496v3-Figure6-1.png': None, '1605.07496v3-Figure7-1.png': None, '1605.07496v3-Figure8-1.png': None, '1605.07496v3-Figure9-1.png': None, '1605.07496v3-Table1-1.png': None, '1605.07496v3-Table2-1.png': None, '1605.07496v3-Table3-1.png': None, '1606.07384v2-Figure1-1.png': None, '1606.07384v2-Figure2-1.png': None, '1608.02784v2-Figure1-1.png': None, '1608.02784v2-Figure2-1.png': None, '1608.02784v2-Figure3-1.png': None, '1608.02784v2-Figure4-1.png': None, '1608.02784v2-Figure5-1.png': None, '1608.02784v2-Figure6-1.png': None, '1608.02784v2-Table1-1.png': None, '1608.02784v2-Table2-1.png': None, '1608.02784v2-Table3-1.png': None, '1611.02654v2-Figure2-1.png': None, '1611.02654v2-Table1-1.png': None, '1611.02654v2-Table2-1.png': None, '1611.02654v2-Table3-1.png': None, '1611.02654v2-Table4-1.png': None, '1611.02654v2-Table5-1.png': None, '1611.03780v2-Figure1-1.png': None, '1611.03780v2-Figure2-1.png': None, '1611.03780v2-Figure3-1.png': None, '1611.03780v2-Figure4-1.png': None, '1611.03780v2-Figure5-1.png': None, '1611.03780v2-Table1-1.png': None, '1611.03780v2-Table2-1.png': None, '1611.03780v2-Table3-1.png': None, '1611.03780v2-Table4-1.png': None, '1611.03780v2-Table5-1.png': None, '1611.04363v2-Figure1-1.png': None, '1611.04363v2-Figure2-1.png': None, '1611.04363v2-Table1-1.png': None, '1611.04684v1-Figure1-1.png': {'caption': 'Architecture of KEHNN', 'content_type': 'figure', 'figure_type': 'schematic'}, '1611.04684v1-Table1-1.png': {'caption': 'A difficult example from QA', 'content_type': 'table', 'figure_type': 'other'}, '1611.04684v1-Table2-1.png': {'caption': 'Table 2: Statistics of the answer selection data set', 'content_type': 'table', 'figure_type': 'N/A'}, '1611.04684v1-Table3-1.png': {'caption': 'Table 3: Evaluation results on answer selection', 'content_type': 'table', 'figure_type': 'N/A'}, '1611.04684v1-Table4-1.png': {'caption': 'Table 4: Evaluation results on response selection', 'content_type': 'table', 'figure_type': 'N/A'}, '1611.04684v1-Table5-1.png': {'caption': 'Accuracy on different length of text', 'content_type': 'table', 'figure_type': 'table'}, '1611.04684v1-Table6-1.png': {'caption': 'Comparison of different channels', 'content_type': 'table', 'figure_type': 'table'}, '1611.05742v3-Figure1-1.png': None, '1611.05742v3-Figure2-1.png': None, '1611.05742v3-Table1-1.png': None, '1611.07718v2-Figure1-1.png': None, '1611.07718v2-Figure2-1.png': None, '1611.07718v2-Figure3-1.png': None, '1611.07718v2-Figure4-1.png': None, '1611.07718v2-Figure5-1.png': None, '1611.07718v2-Figure6-1.png': None, '1611.07718v2-Figure7-1.png': None, '1611.07718v2-Figure8-1.png': None, '1611.07718v2-Table2-1.png': None, '1611.07718v2-Table3-1.png': None, '1611.07718v2-Table4-1.png': None, '1611.07718v2-Table5-1.png': None, '1612.02803v5-Figure1-1.png': None, '1612.02803v5-Figure2-1.png': None, '1612.02803v5-Table1-1.png': None, '1701.03077v10-Figure1-1.png': None, '1701.03077v10-Figure10-1.png': None, '1701.03077v10-Figure11-1.png': None, '1701.03077v10-Figure12-1.png': None, '1701.03077v10-Figure13-1.png': None, '1701.03077v10-Figure14-1.png': None, '1701.03077v10-Figure15-1.png': None, '1701.03077v10-Figure16-1.png': None, '1701.03077v10-Figure17-1.png': None, '1701.03077v10-Figure2-1.png': None, '1701.03077v10-Figure3-1.png': None, '1701.03077v10-Figure4-1.png': None, '1701.03077v10-Figure5-1.png': None, '1701.03077v10-Figure6-1.png': None, '1701.03077v10-Figure7-1.png': None, '1701.03077v10-Figure8-1.png': None, '1701.03077v10-Figure9-1.png': None, '1701.03077v10-Table1-1.png': None, '1701.03077v10-Table2-1.png': None, '1701.03077v10-Table3-1.png': None, '1701.03077v10-Table4-1.png': None, '1701.06171v4-Figure1-1.png': None, '1701.06171v4-Figure2-1.png': None, '1701.06171v4-Figure3-1.png': None, '1701.06171v4-Figure4-1.png': None, '1701.06171v4-Figure5-1.png': None, '1701.06171v4-Table1-1.png': None, '1701.06171v4-Table2-1.png': None, '1702.03584v3-Figure1-1.png': None, '1702.03584v3-Figure2-1.png': None, '1702.03584v3-Table1-1.png': None, '1702.08694v3-Figure1-1.png': None, '1702.08694v3-Figure2-1.png': None, '1702.08694v3-Figure3-1.png': None, '1702.08694v3-Figure4-1.png': None, '1702.08694v3-Figure5-1.png': None, '1702.08694v3-Table1-1.png': None, '1702.08694v3-Table2-1.png': None, '1703.00060v2-Table1-1.png': None, '1703.00060v2-Table2-1.png': None, '1703.00899v2-Figure1-1.png': None, '1703.02507v3-Figure1-1.png': None, '1703.02507v3-Table1-1.png': None, '1703.02507v3-Table2-1.png': None, '1703.02507v3-Table3-1.png': None, '1703.02507v3-Table4-1.png': None, '1703.02507v3-Table5-1.png': None, '1703.02507v3-Table6-1.png': None, '1703.02507v3-Table7-1.png': None, '1703.02507v3-Table8-1.png': None, '1703.04887v4-Figure1-1.png': None, '1703.04887v4-Figure2-1.png': None, '1703.04887v4-Table1-1.png': None, '1703.04887v4-Table2-1.png': None, '1703.04887v4-Table3-1.png': None, '1703.07015v3-Figure1-1.png': None, '1703.07015v3-Figure2-1.png': None, '1703.07015v3-Figure3-1.png': None, '1703.07015v3-Figure4-1.png': None, '1703.07015v3-Figure5-1.png': None, '1703.07015v3-Figure56-1.png': None, '1703.07015v3-Figure6-1.png': None, '1703.07015v3-Figure7-1.png': None, '1703.07015v3-Table1-1.png': None, '1703.07015v3-Table2-1.png': None, '1703.10730v2-Figure1-1.png': None, '1703.10730v2-Figure10-1.png': None, '1703.10730v2-Figure11-1.png': None, '1703.10730v2-Figure12-1.png': None, '1703.10730v2-Figure13-1.png': None, '1703.10730v2-Figure14-1.png': None, '1703.10730v2-Figure2-1.png': None, '1703.10730v2-Figure3-1.png': None, '1703.10730v2-Figure4-1.png': None, '1703.10730v2-Figure6-1.png': None, '1703.10730v2-Figure7-1.png': None, '1703.10730v2-Figure8-1.png': None, '1703.10730v2-Figure9-1.png': None, '1704.00774v3-Figure1-1.png': None, '1704.00774v3-Table1-1.png': None, '1704.04539v2-Figure3-1.png': None, '1704.04539v2-Figure4-1.png': None, '1704.04539v2-Figure5-1.png': None, '1704.04539v2-Table1-1.png': None, '1704.04539v2-Table2-1.png': None, '1704.05426v4-Figure1-1.png': None, '1704.05426v4-Table1-1.png': None, '1704.05426v4-Table2-1.png': None, '1704.05426v4-Table3-1.png': None, '1704.05426v4-Table4-1.png': None, '1704.05426v4-Table5-1.png': None, '1704.05958v2-Figure1-1.png': None, '1704.05958v2-Figure2-1.png': None, '1704.05958v2-Figure3-1.png': None, '1704.05958v2-Figure4-1.png': None, '1704.05958v2-Figure5-1.png': None, '1704.05958v2-Figure6-1.png': None, '1704.05958v2-Figure7-1.png': None, '1704.05958v2-Table1-1.png': None, '1704.05958v2-Table2-1.png': None, '1704.05958v2-Table3-1.png': None, '1704.07121v2-Figure1-1.png': None, '1704.07121v2-Figure2-1.png': None, '1704.07121v2-Figure3-1.png': None, '1704.07121v2-Figure4-1.png': None, '1704.07121v2-Figure5-1.png': None, '1704.07121v2-Figure6-1.png': None, '1704.07121v2-Figure7-1.png': None, '1704.07121v2-Figure8-1.png': None, '1704.07121v2-Table1-1.png': None, '1704.07121v2-Table10-1.png': None, '1704.07121v2-Table11-1.png': None, '1704.07121v2-Table12-1.png': None, '1704.07121v2-Table2-1.png': None, '1704.07121v2-Table3-1.png': None, '1704.07121v2-Table4-1.png': None, '1704.07121v2-Table5-1.png': None, '1704.07121v2-Table6-1.png': None, '1704.07121v2-Table7-1.png': None, '1704.07121v2-Table8-1.png': None, '1704.07121v2-Table9-1.png': None, '1704.07854v4-Figure1-1.png': None, '1704.07854v4-Figure10-1.png': None, '1704.07854v4-Figure11-1.png': None, '1704.07854v4-Figure12-1.png': None, '1704.07854v4-Figure13-1.png': None, '1704.07854v4-Figure14-1.png': None, '1704.07854v4-Figure15-1.png': None, '1704.07854v4-Figure16-1.png': None, '1704.07854v4-Figure17-1.png': None, '1704.07854v4-Figure18-1.png': None, '1704.07854v4-Figure2-1.png': None, '1704.07854v4-Figure3-1.png': None, '1704.07854v4-Figure4-1.png': None, '1704.07854v4-Figure5-1.png': None, '1704.07854v4-Figure6-1.png': None, '1704.07854v4-Figure7-1.png': None, '1704.07854v4-Figure8-1.png': None, '1704.07854v4-Figure9-1.png': None, '1704.07854v4-Table1-1.png': None, '1704.07854v4-Table2-1.png': None, '1704.08615v2-Figure1-1.png': None, '1704.08615v2-Figure2-1.png': None, '1704.08615v2-Figure3-1.png': None, '1704.08615v2-Figure4-1.png': None, '1704.08615v2-Figure5-1.png': None, '1704.08615v2-Figure6-1.png': None, '1704.08615v2-Figure7-1.png': None, '1704.08615v2-Table1-1.png': None, '1704.08615v2-Table2-1.png': None, '1704.08615v2-Table3-1.png': None, '1705.02798v6-Figure1-1.png': None, '1705.02798v6-Figure2-1.png': None, '1705.02798v6-Figure3-1.png': None, '1705.02798v6-Figure4-1.png': None, '1705.02798v6-Figure5-1.png': None, '1705.02798v6-Table1-1.png': None, '1705.02798v6-Table2-1.png': None, '1705.02798v6-Table3-1.png': None, '1705.02798v6-Table4-1.png': None, '1705.02798v6-Table5-1.png': None, '1705.02946v3-Figure1-1.png': None, '1705.02946v3-Figure2-1.png': None, '1705.02946v3-Figure3-1.png': None, '1705.02946v3-Figure4-1.png': None, '1705.02946v3-Figure5-1.png': None, '1705.02946v3-Figure6-1.png': None, '1705.02946v3-Figure7-1.png': None, '1705.02946v3-Figure8-1.png': None, '1705.02946v3-Table1-1.png': None, '1705.02946v3-Table2-1.png': None, '1705.07164v8-Figure1-1.png': None, '1705.07164v8-Figure3-1.png': None, '1705.07164v8-Table1-1.png': None, '1705.07164v8-Table2-1.png': None, '1705.07384v2-Figure2-1.png': None, '1705.07384v2-Table2-1.png': None, '1705.08016v3-Figure1-1.png': None, '1705.08016v3-Figure2-1.png': None, '1705.08016v3-Figure3-1.png': None, '1705.08016v3-Table1-1.png': None, '1705.08016v3-Table2-1.png': None, '1705.08016v3-Table3-1.png': None, '1705.08016v3-Table4-1.png': None, '1705.09296v2-Figure1-1.png': None, '1705.09296v2-Figure2-1.png': None, '1705.09296v2-Figure3-1.png': None, '1705.09296v2-Table1-1.png': None, '1705.09296v2-Table2-1.png': None, '1705.09296v2-Table3-1.png': None, '1705.09296v2-Table4-1.png': None, '1705.09296v2-Table5-1.png': None, '1705.09296v2-Table6-1.png': None, '1705.09882v2-Figure1-1.png': None, '1705.09882v2-Figure2-1.png': None, '1705.09882v2-Figure3-1.png': None, '1705.09882v2-Figure4-1.png': None, '1705.09882v2-Figure5-1.png': None, '1705.09882v2-Figure6-1.png': None, '1705.09882v2-Figure7-1.png': None, '1705.09882v2-Table1-1.png': None, '1705.09882v2-Table2-1.png': None, '1705.09966v2-Figure1-1.png': None, '1705.09966v2-Figure10-1.png': None, '1705.09966v2-Figure11-1.png': None, '1705.09966v2-Figure12-1.png': None, '1705.09966v2-Figure13-1.png': None, '1705.09966v2-Figure14-1.png': None, '1705.09966v2-Figure15-1.png': None, '1705.09966v2-Figure2-1.png': None, '1705.09966v2-Figure3-1.png': None, '1705.09966v2-Figure4-1.png': None, '1705.09966v2-Figure5-1.png': None, '1705.09966v2-Figure6-1.png': None, '1705.09966v2-Figure7-1.png': None, '1705.09966v2-Figure8-1.png': None, '1705.09966v2-Figure9-1.png': None, '1705.09966v2-Table1-1.png': None, '1705.10667v4-Figure1-1.png': None, '1705.10667v4-Figure2-1.png': None, '1705.10667v4-Figure3-1.png': None, '1705.10667v4-Table1-1.png': None, '1705.10667v4-Table2-1.png': None, '1705.10667v4-Table3-1.png': None, '1705.10667v4-Table4-1.png': None, '1705.10667v4-Table5-1.png': None, '1706.00633v4-Figure1-1.png': None, '1706.00633v4-Figure2-1.png': None, '1706.00633v4-Figure3-1.png': None, '1706.00633v4-Figure4-1.png': None, '1706.00633v4-Figure5-1.png': None, '1706.00633v4-Figure6-1.png': None, '1706.00633v4-Figure7-1.png': None, '1706.00633v4-Table1-1.png': None, '1706.00633v4-Table2-1.png': None, '1706.00633v4-Table3-1.png': None, '1706.00633v4-Table4-1.png': None, '1706.00633v4-Table5-1.png': None, '1706.00633v4-Table6-1.png': None, '1706.00827v2-Figure1-1.png': None, '1706.00827v2-Figure2-1.png': None, '1706.00827v2-Figure3-1.png': None, '1706.00827v2-Figure4-1.png': None, '1706.00827v2-Figure5-1.png': None, '1706.00827v2-Figure6-1.png': None, '1706.00827v2-Table1-1.png': None, '1706.00827v2-Table2-1.png': None, '1706.00827v2-Table3-1.png': None, '1706.00827v2-Table4-1.png': None, '1706.00827v2-Table5-1.png': None, '1706.00827v2-Table6-1.png': None, '1706.00827v2-Table7-1.png': None, '1706.00827v2-Table8-1.png': None, '1706.03847v3-Figure1-1.png': None, '1706.03847v3-Figure2-1.png': None, '1706.03847v3-Figure3-1.png': None, '1706.03847v3-Figure4-1.png': None, '1706.03847v3-Figure5-1.png': None, '1706.03847v3-Figure6-1.png': None, '1706.03847v3-Table1-1.png': None, '1706.03847v3-Table2-1.png': None, '1706.03847v3-Table3-1.png': None, '1706.04269v2-Figure1-1.png': None, '1706.04269v2-Figure2-1.png': None, '1706.04269v2-Figure3-1.png': None, '1706.04269v2-Figure4-1.png': None, '1706.04269v2-Figure5-1.png': None, '1706.04269v2-Table1-1.png': None, '1706.04284v3-Figure1-1.png': None, '1706.04284v3-Figure2-1.png': None, '1706.04284v3-Figure3-1.png': None, '1706.04284v3-Figure4-1.png': None, '1706.04284v3-Figure5-1.png': None, '1706.04284v3-Table1-1.png': None, '1706.04284v3-Table2-1.png': None, '1706.04284v3-Table3-1.png': None, '1706.08146v3-Figure1-1.png': None, '1706.08146v3-Figure2-1.png': None, '1706.08146v3-Figure3-1.png': None, '1706.08146v3-Figure4-1.png': None, '1706.08146v3-Figure5-1.png': None, '1706.08146v3-Figure6-1.png': None, '1706.08146v3-Table1-1.png': None, '1707.00189v3-Table1-1.png': None, '1707.00189v3-Table2-1.png': None, '1707.00524v2-Figure1-1.png': None, '1707.00524v2-Figure2-1.png': None, '1707.00524v2-Figure3-1.png': None, '1707.00524v2-Figure4-1.png': None, '1707.00524v2-Figure5-1.png': None, '1707.00524v2-Table1-1.png': None, '1707.00524v2-Table2-1.png': None, '1707.01917v2-Figure1-1.png': None, '1707.01917v2-Figure2-1.png': None, '1707.01917v2-Table1-1.png': None, '1707.01917v2-Table2-1.png': None, '1707.01917v2-Table3-1.png': None, '1707.01917v2-Table4-1.png': None, '1707.01917v2-Table5-1.png': None, '1707.01922v5-Figure1-1.png': None, '1707.01922v5-Figure2-1.png': None, '1707.01922v5-Figure3-1.png': None, '1707.01922v5-Figure4-1.png': None, '1707.01922v5-Table1-1.png': None, '1707.01922v5-Table2-1.png': None, '1707.01922v5-Table3-1.png': None, '1707.01922v5-Table4-1.png': None, '1707.01922v5-Table5-1.png': None, '1707.01922v5-Table6-1.png': None, '1707.01922v5-Table7-1.png': None, '1707.01922v5-Table8-1.png': None, '1707.01922v5-Table9-1.png': None, '1707.06320v2-Figure1-1.png': None, '1707.06320v2-Table1-1.png': None, '1707.06320v2-Table2-1.png': None, '1707.06320v2-Table3-1.png': None, '1707.06320v2-Table4-1.png': None, '1707.06320v2-Table5-1.png': None, '1707.08608v3-Table1-1.png': None, '1707.08608v3-Table10-1.png': None, '1707.08608v3-Table11-1.png': None, '1707.08608v3-Table2-1.png': None, '1707.08608v3-Table3-1.png': None, '1707.08608v3-Table4-1.png': None, '1707.08608v3-Table5-1.png': None, '1707.08608v3-Table6-1.png': None, '1707.08608v3-Table7-1.png': None, '1707.08608v3-Table8-1.png': None, '1707.08608v3-Table9-1.png': None, '1708.00160v2-Figure1-1.png': None, '1708.00160v2-Figure2-1.png': None, '1708.00160v2-Figure3-1.png': None, '1708.00160v2-Figure4-1.png': None, '1708.00160v2-Figure5-1.png': None, '1708.00160v2-Table1-1.png': None, '1708.00160v2-Table2-1.png': None, '1708.00160v2-Table3-1.png': None, '1708.00160v2-Table4-1.png': None, '1708.00160v2-Table5-1.png': None, '1708.00160v2-Table6-1.png': None, '1708.00160v2-Table7-1.png': None, '1708.01425v4-Figure2-1.png': None, '1708.01425v4-Figure3-1.png': None, '1708.01425v4-Figure4-1.png': None, '1708.01425v4-Figure5-1.png': None, '1708.01425v4-Table1-1.png': None, '1708.01425v4-Table2-1.png': None, '1708.02153v2-Figure1-1.png': None, '1708.02153v2-Table1-1.png': None, '1708.02153v2-Table2-1.png': None, '1708.02153v2-Table3-1.png': None, '1708.02153v2-Table4-1.png': None, '1708.03797v1-Figure1-1.png': None, '1708.03797v1-Table1-1.png': None, '1708.03797v1-Table2-1.png': None, '1708.05239v3-Figure1-1.png': None, '1708.05239v3-Figure10-1.png': None, '1708.05239v3-Figure2-1.png': None, '1708.05239v3-Figure3-1.png': None, '1708.05239v3-Figure4-1.png': None, '1708.05239v3-Figure5-1.png': None, '1708.05239v3-Figure6-1.png': None, '1708.05239v3-Figure7-1.png': None, '1708.05239v3-Figure8-1.png': None, '1708.05239v3-Figure9-1.png': None, '1708.06832v3-Figure1-1.png': None, '1708.06832v3-Figure2-1.png': None, '1708.06832v3-Figure3-1.png': None, '1708.06832v3-Figure4-1.png': None, '1708.06832v3-Figure5-1.png': None, '1709.00139v4-Figure1-1.png': None, '1709.00139v4-Table1-1.png': None, '1709.02418v2-Figure1-1.png': None, '1709.02418v2-Figure2-1.png': None, '1709.02755v5-Figure1-1.png': None, '1709.02755v5-Figure2-1.png': None, '1709.02755v5-Figure3-1.png': None, '1709.02755v5-Figure6-1.png': None, '1709.02755v5-Table1-1.png': None, '1709.02755v5-Table2-1.png': None, '1709.02755v5-Table4-1.png': None, '1709.02755v5-Table6-1.png': None, '1709.08294v3-Figure1-1.png': None, '1709.08294v3-Figure2-1.png': None, '1709.08294v3-Figure3-1.png': None, '1709.08294v3-Table1-1.png': None, '1709.08294v3-Table2-1.png': None, '1709.08294v3-Table3-1.png': None, '1709.08294v3-Table4-1.png': None, '1709.08294v3-Table5-1.png': None, '1710.01507v4-Figure1-1.png': None, '1710.01507v4-Table1-1.png': None, '1710.05654v2-Figure1-1.png': None, '1710.05654v2-Figure10-1.png': None, '1710.05654v2-Figure11-1.png': None, '1710.05654v2-Figure12-1.png': None, '1710.05654v2-Figure13-1.png': None, '1710.05654v2-Figure14-1.png': None, '1710.05654v2-Figure15-1.png': None, '1710.05654v2-Figure4-1.png': None, '1710.05654v2-Figure5-1.png': None, '1710.05654v2-Figure6-1.png': None, '1710.05654v2-Figure7-1.png': None, '1710.05654v2-Figure8-1.png': None, '1710.05654v2-Figure9-1.png': None, '1710.05654v2-Table1-1.png': None, '1710.06177v2-Figure1-1.png': None, '1710.06177v2-Figure2-1.png': None, '1710.06177v2-Figure3-1.png': None, '1710.06177v2-Figure4-1.png': None, '1710.06177v2-Figure5-1.png': None, '1710.06177v2-Table1-1.png': None, '1710.06177v2-Table2-1.png': None, '1710.06177v2-Table3-1.png': None, '1802.07222v1-Figure1-1.png': None, '1802.07222v1-Figure10-1.png': None, '1802.07222v1-Figure11-1.png': None, '1802.07222v1-Figure12-1.png': None, '1802.07222v1-Figure13-1.png': None, '1802.07222v1-Figure14-1.png': None, '1802.07222v1-Figure15-1.png': None, '1802.07222v1-Figure16-1.png': None, '1802.07222v1-Figure2-1.png': None, '1802.07222v1-Figure3-1.png': None, '1802.07222v1-Figure4-1.png': None, '1802.07222v1-Figure5-1.png': None, '1802.07222v1-Figure6-1.png': None, '1802.07222v1-Figure7-1.png': None, '1802.07222v1-Figure8-1.png': None, '1802.07222v1-Figure9-1.png': None, '1802.07222v1-Table1-1.png': None, '1802.07222v1-Table2-1.png': None, '1802.07351v2-Figure1-1.png': None, '1802.07351v2-Figure10-1.png': None, '1802.07351v2-Figure2-1.png': None, '1802.07351v2-Figure3-1.png': None, '1802.07351v2-Figure4-1.png': None, '1802.07351v2-Figure5-1.png': None, '1802.07351v2-Figure6-1.png': None, '1802.07351v2-Figure7-1.png': None, '1802.07351v2-Figure8-1.png': None, '1802.07351v2-Figure9-1.png': None, '1802.07351v2-Table1-1.png': None, '1802.07351v2-Table2-1.png': None, '1802.07351v2-Table3-1.png': None, '1802.07351v2-Table4-1.png': None, '1802.07351v2-Table5-1.png': None, '1802.07351v2-Table6-1.png': None, '1802.07459v2-Figure1-1.png': None, '1802.07459v2-Figure2-1.png': None, '1802.07459v2-Figure3-1.png': None, '1802.07459v2-Table1-1.png': None, '1802.07459v2-Table2-1.png': None, '1803.01128v3-Table1-1.png': None, '1803.01128v3-Table2-1.png': None, '1803.01128v3-Table3-1.png': None, '1803.01128v3-Table4-1.png': None, '1803.01128v3-Table5-1.png': None, '1803.01128v3-Table6-1.png': None, '1803.01128v3-Table7-1.png': None, '1803.01128v3-Table8-1.png': None, '1803.01128v3-Table9-1.png': None, '1803.02750v3-Figure1-1.png': None, '1803.02750v3-Figure10-1.png': None, '1803.02750v3-Figure12-1.png': None, '1803.02750v3-Figure13-1.png': None, '1803.02750v3-Figure2-1.png': None, '1803.02750v3-Figure3-1.png': None, '1803.02750v3-Figure4-1.png': None, '1803.02750v3-Figure5-1.png': None, '1803.02750v3-Figure6-1.png': None, '1803.02750v3-Figure7-1.png': None, '1803.02750v3-Figure8-1.png': None, '1803.02750v3-Figure9-1.png': None, '1803.02750v3-TableI-1.png': None, '1803.02750v3-TableII-1.png': None, '1803.02750v3-TableIII-1.png': None, '1803.02750v3-TableIV-1.png': None, '1803.03467v4-Figure1-1.png': None, '1803.03467v4-Figure2-1.png': None, '1803.03467v4-Figure3-1.png': None, '1803.03467v4-Figure4-1.png': None, '1803.03467v4-Figure5-1.png': None, '1803.03467v4-Figure6-1.png': None, '1803.03467v4-Figure7-1.png': None, '1803.03467v4-Figure8-1.png': None, '1803.03467v4-Figure9-1.png': None, '1803.03467v4-Table1-1.png': None, '1803.03467v4-Table2-1.png': None, '1803.03467v4-Table3-1.png': None, '1803.03467v4-Table4-1.png': None, '1803.03467v4-Table5-1.png': None, '1803.04383v2-Figure1-1.png': None, '1803.04383v2-Figure2-1.png': None, '1803.04383v2-Figure3-1.png': None, '1803.04383v2-Figure4-1.png': None, '1803.04383v2-Figure5-1.png': None, '1803.04383v2-Figure6-1.png': None, '1803.04572v2-Figure1-1.png': None, '1803.04572v2-Figure2-1.png': None, '1803.04572v2-Figure3-1.png': None, '1803.04572v2-Figure4-1.png': None, '1803.04572v2-Figure5-1.png': None, '1803.04572v2-Figure6-1.png': None, '1803.04572v2-Figure7-1.png': None, '1803.04572v2-Figure8-1.png': None, '1803.04572v2-Table1-1.png': None, '1803.04572v2-Table2-1.png': None, '1803.04572v2-Table3-1.png': None, '1803.04572v2-Table4-1.png': None, '1803.04572v2-Table5-1.png': None, '1803.04572v2-Table6-1.png': None, '1803.04572v2-Table7-1.png': None, '1803.05776v2-Figure1-1.png': None, '1803.06506v3-Figure1-1.png': None, '1803.06506v3-Figure2-1.png': None, '1803.06506v3-Figure3-1.png': None, '1803.06506v3-Figure4-1.png': None, '1803.06506v3-Figure5-1.png': None, '1803.06506v3-Figure6-1.png': None, '1803.06506v3-Figure7-1.png': None, '1803.06506v3-Table1-1.png': None, '1803.06506v3-Table2-1.png': None, '1803.06506v3-Table3-1.png': None, '1804.00863v3-Figure1-1.png': None, '1804.00863v3-Figure10-1.png': None, '1804.00863v3-Figure11-1.png': None, '1804.00863v3-Figure12-1.png': None, '1804.00863v3-Figure2-1.png': None, '1804.00863v3-Figure3-1.png': None, '1804.00863v3-Figure4-1.png': None, '1804.00863v3-Figure5-1.png': None, '1804.00863v3-Figure6-1.png': None, '1804.00863v3-Figure7-1.png': None, '1804.00863v3-Figure8-1.png': None, '1804.00863v3-Figure9-1.png': None, '1804.00863v3-Table1-1.png': None, '1804.01429v3-Figure1-1.png': None, '1804.01429v3-Figure10-1.png': None, '1804.01429v3-Figure2-1.png': None, '1804.01429v3-Figure3-1.png': None, '1804.01429v3-Figure4-1.png': None, '1804.01429v3-Figure5-1.png': None, '1804.01429v3-Figure6-1.png': None, '1804.01429v3-Figure7-1.png': None, '1804.01429v3-Figure8-1.png': None, '1804.01429v3-Figure9-1.png': None, '1804.01429v3-Table1-1.png': None, '1804.04410v2-Figure1-1.png': None, '1804.04410v2-Figure2-1.png': None, '1804.04410v2-Table1-1.png': None, '1804.04786v3-Figure1-1.png': None, '1804.04786v3-Figure2-1.png': None, '1804.04786v3-Figure3-1.png': None, '1804.04786v3-Figure4-1.png': None, '1804.04786v3-Figure5-1.png': None, '1804.04786v3-Figure6-1.png': None, '1804.04786v3-Table1-1.png': None, '1804.05936v2-Figure1-1.png': None, '1804.05936v2-Figure2-1.png': None, '1804.05936v2-Figure3-1.png': None, '1804.05936v2-Figure4-1.png': None, '1804.05936v2-Table1-1.png': None, '1804.05936v2-Table2-1.png': None, '1804.05936v2-Table3-1.png': None, '1804.05936v2-Table4-1.png': None, '1804.05936v2-Table5-1.png': None, '1804.05938v2-Figure1-1.png': None, '1804.05938v2-Figure2-1.png': None, '1804.05938v2-Table1-1.png': None, '1804.05938v2-Table2-1.png': None, '1804.05938v2-Table3-1.png': None, '1804.05938v2-Table4-1.png': None, '1804.05995v2-Figure3-1.png': None, '1804.05995v2-Figure4-1.png': None, '1804.05995v2-Figure5-1.png': None, '1804.05995v2-Table1-1.png': None, '1804.07707v2-Figure1-1.png': None, '1804.07707v2-Table1-1.png': None, '1804.07707v2-Table2-1.png': None, '1804.07707v2-Table3-1.png': None, '1804.07849v4-Figure1-1.png': None, '1804.07849v4-Table1-1.png': None, '1804.07849v4-Table2-1.png': None, '1804.07849v4-Table3-1.png': None, '1804.07849v4-Table4-1.png': None, '1804.07931v2-Figure1-1.png': None, '1804.07931v2-Figure2-1.png': None, '1804.07931v2-Figure3-1.png': None, '1804.07931v2-Table1-1.png': None, '1804.07931v2-Table2-1.png': None, '1805.00912v4-Figure1-1.png': None, '1805.00912v4-Figure2-1.png': None, '1805.00912v4-Figure3-1.png': None, '1805.00912v4-Table1-1.png': None, '1805.00912v4-Table2-1.png': None, '1805.00912v4-Table3-1.png': None, '1805.00912v4-Table4-1.png': None, '1805.00912v4-Table5-1.png': None, '1805.00912v4-Table6-1.png': None, '1805.01216v3-Figure1-1.png': None, '1805.01216v3-Figure10-1.png': None, '1805.01216v3-Figure2-1.png': None, '1805.01216v3-Figure3-1.png': None, '1805.01216v3-Figure4-1.png': None, '1805.01216v3-Figure5-1.png': None, '1805.01216v3-Figure7-1.png': None, '1805.01216v3-Figure8-1.png': None, '1805.01216v3-Figure9-1.png': None, '1805.01216v3-Table1-1.png': None, '1805.01216v3-Table10-1.png': None, '1805.01216v3-Table11-1.png': None, '1805.01216v3-Table12-1.png': None, '1805.01216v3-Table13-1.png': None, '1805.01216v3-Table14-1.png': None, '1805.01216v3-Table2-1.png': None, '1805.01216v3-Table3-1.png': None, '1805.01216v3-Table4-1.png': None, '1805.01216v3-Table5-1.png': None, '1805.01216v3-Table6-1.png': None, '1805.01216v3-Table7-1.png': None, '1805.01216v3-Table8-1.png': None, '1805.01216v3-Table9-1.png': None, '1805.02349v2-Figure1-1.png': None, '1805.02349v2-Figure2-1.png': None, '1805.04609v3-Figure1-1.png': None, '1805.04609v3-Figure2-1.png': None, '1805.04609v3-Figure3-1.png': None, '1805.04687v2-Figure1-1.png': None, '1805.04687v2-Figure10-1.png': None, '1805.04687v2-Figure11-1.png': None, '1805.04687v2-Figure12-1.png': None, '1805.04687v2-Figure13-1.png': None, '1805.04687v2-Figure14-1.png': None, '1805.04687v2-Figure15-1.png': None, '1805.04687v2-Figure2-1.png': None, '1805.04687v2-Figure3-1.png': None, '1805.04687v2-Figure4-1.png': None, '1805.04687v2-Figure5-1.png': None, '1805.04687v2-Figure6-1.png': None, '1805.04687v2-Figure7-1.png': None, '1805.04687v2-Figure8-1.png': None, '1805.04687v2-Figure9-1.png': None, '1805.04687v2-Table1-1.png': None, '1805.04687v2-Table10-1.png': None, '1805.04687v2-Table11-1.png': None, '1805.04687v2-Table12-1.png': None, '1805.04687v2-Table13-1.png': None, '1805.04687v2-Table14-1.png': None, '1805.04687v2-Table2-1.png': None, '1805.04687v2-Table3-1.png': None, '1805.04687v2-Table4-1.png': None, '1805.04687v2-Table5-1.png': None, '1805.04687v2-Table6-1.png': None, '1805.04687v2-Table7-1.png': None, '1805.04687v2-Table8-1.png': None, '1805.04687v2-Table9-1.png': None, '1805.06431v4-Figure1-1.png': None, '1805.06431v4-Figure10-1.png': None, '1805.06431v4-Figure12-1.png': None, '1805.06431v4-Figure13-1.png': None, '1805.06431v4-Figure14-1.png': None, '1805.06431v4-Figure2-1.png': None, '1805.06431v4-Figure3-1.png': None, '1805.06431v4-Figure4-1.png': None, '1805.06431v4-Figure5-1.png': None, '1805.06431v4-Figure6-1.png': None, '1805.06431v4-Figure7-1.png': None, '1805.06431v4-Figure8-1.png': None, '1805.06431v4-Figure9-1.png': None, '1805.06431v4-Table1-1.png': None, '1805.06431v4-Table10-1.png': None, '1805.06431v4-Table11-1.png': None, '1805.06431v4-Table12-1.png': None, '1805.06431v4-Table13-1.png': None, '1805.06431v4-Table14-1.png': None, '1805.06431v4-Table15-1.png': None, '1805.06431v4-Table2-1.png': None, '1805.06431v4-Table3-1.png': None, '1805.06431v4-Table4-1.png': None, '1805.06431v4-Table5-1.png': None, '1805.06431v4-Table6-1.png': None, '1805.06431v4-Table7-1.png': None, '1805.06431v4-Table8-1.png': None, '1805.06431v4-Table9-1.png': None, '1805.06447v3-Figure1-1.png': None, '1805.06447v3-Figure2-1.png': None, '1805.06447v3-Figure3-1.png': None, '1805.06447v3-Figure4-1.png': None, '1805.06447v3-Figure5-1.png': None, '1805.06447v3-Table1-1.png': None, '1805.06447v3-Table2-1.png': None, '1805.06447v3-Table3-1.png': None, '1805.06447v3-Table4-1.png': None, '1805.06447v3-Table5-1.png': None, '1805.06447v3-Table6-1.png': None, '1805.06447v3-Table7-1.png': None, '1805.06447v3-Table8-1.png': None, '1805.07567v2-Figure1-1.png': None, '1805.07567v2-Figure2-1.png': None, '1805.07567v2-Figure3-1.png': None, '1805.07567v2-Figure4-1.png': None, '1805.07567v2-Figure5-1.png': None, '1805.07567v2-Figure6-1.png': None, '1805.07567v2-Figure7-1.png': None, '1805.07567v2-Table1-1.png': None, '1805.07567v2-Table2-1.png': None, '1805.07567v2-Table3-1.png': None, '1805.07567v2-Table4-1.png': None, '1805.08465v3-Figure1-1.png': None, '1805.08465v3-Figure2-1.png': None, '1805.08465v3-Figure3-1.png': None, '1805.08465v3-Table1-1.png': None, '1805.08751v2-Figure2-1.png': None, '1805.08751v2-Figure3-1.png': None, '1805.08751v2-Figure4-1.png': None, '1805.08751v2-Figure5-1.png': None, '1805.08751v2-Figure6-1.png': None, '1805.08751v2-Figure7-1.png': None, '1805.08751v2-TableI-1.png': None, '1809.00263v5-Figure1-1.png': None, '1809.00263v5-Figure10-1.png': None, '1809.00263v5-Figure11-1.png': None, '1809.00263v5-Figure12-1.png': None, '1809.00263v5-Figure13-1.png': None, '1809.00263v5-Figure14-1.png': None, '1809.00263v5-Figure15-1.png': None, '1809.00263v5-Figure16-1.png': None, '1809.00263v5-Figure2-1.png': None, '1809.00263v5-Figure3-1.png': None, '1809.00263v5-Figure4-1.png': None, '1809.00263v5-Figure5-1.png': None, '1809.00263v5-Figure6-1.png': None, '1809.00263v5-Figure7-1.png': None, '1809.00263v5-Figure9-1.png': None, '1809.00263v5-Table1-1.png': None, '1809.00263v5-Table2-1.png': None, '1809.00263v5-Table3-1.png': None, '1809.00458v1-Figure1-1.png': None, '1809.00458v1-Figure10-1.png': None, '1809.00458v1-Figure11-1.png': None, '1809.00458v1-Figure12-1.png': None, '1809.00458v1-Figure13-1.png': None, '1809.00458v1-Figure14-1.png': None, '1809.00458v1-Figure15-1.png': None, '1809.00458v1-Figure16-1.png': None, '1809.00458v1-Figure17-1.png': None, '1809.00458v1-Figure18-1.png': None, '1809.00458v1-Figure19-1.png': None, '1809.00458v1-Figure2-1.png': None, '1809.00458v1-Figure3-1.png': None, '1809.00458v1-Figure4-1.png': None, '1809.00458v1-Figure5-1.png': None, '1809.00458v1-Figure7-1.png': None, '1809.00458v1-Figure8-1.png': None, '1809.00458v1-Figure9-1.png': None, '1809.00458v1-TableI-1.png': None, '1809.00458v1-TableII-1.png': None, '1809.00458v1-TableIII-1.png': None, '1809.01246v1-Figure1-1.png': None, '1809.01246v1-Figure10-1.png': None, '1809.01246v1-Figure11-1.png': None, '1809.01246v1-Figure12-1.png': None, '1809.01246v1-Figure13-1.png': None, '1809.01246v1-Figure2-1.png': None, '1809.01246v1-Figure3-1.png': None, '1809.01246v1-Figure4-1.png': None, '1809.01246v1-Figure5-1.png': None, '1809.01246v1-Figure6-1.png': None, '1809.01246v1-Figure8-1.png': None, '1809.01246v1-Figure9-1.png': None, '1809.01246v1-TableI-1.png': None, '1809.01989v2-Figure1-1.png': None, '1809.01989v2-Table1-1.png': None, '1809.02731v3-Table1-1.png': None, '1809.02731v3-Table2-1.png': None, '1809.02731v3-Table3-1.png': None, '1809.02731v3-Table4-1.png': None, '1809.02731v3-Table5-1.png': None, '1809.03149v2-Figure1-1.png': None, '1809.03149v2-Figure2-1.png': None, '1809.03149v2-Figure3-1.png': None, '1809.03149v2-Figure4-1.png': None, '1809.03149v2-Figure5-1.png': None, '1809.03149v2-Figure6-1.png': None, '1809.03149v2-Figure7-1.png': None, '1809.03149v2-Figure8-1.png': None, '1809.03149v2-Figure9-1.png': None, '1809.03149v2-Table1-1.png': None, '1809.03149v2-Table2-1.png': None, '1809.03149v2-Table3-1.png': None, '1809.03449v3-Figure1-1.png': None, '1809.03449v3-Figure2-1.png': None, '1809.03449v3-Figure3-1.png': None, '1809.03449v3-Figure4-1.png': None, '1809.03449v3-Table1-1.png': None, '1809.03449v3-Table2-1.png': None, '1809.03449v3-Table3-1.png': None, '1809.03550v3-Figure1-1.png': None, '1809.03550v3-Figure2-1.png': None, '1809.03550v3-Figure3-1.png': None, '1809.03550v3-Figure4-1.png': None, '1809.03550v3-Table1-1.png': None, '1809.03550v3-Table2-1.png': None, '1809.03550v3-Table3-1.png': None, '1809.03550v3-Table4-1.png': None, '1809.03550v3-Table5-1.png': None, '1809.03550v3-Table6-1.png': None, '1809.04276v2-Figure1-1.png': None, '1809.04276v2-Figure2-1.png': None, '1809.04276v2-Table1-1.png': None, '1809.04276v2-Table2-1.png': None, '1809.04276v2-Table3-1.png': None, '1809.04276v2-Table4-1.png': None, '1809.04276v2-Table5-1.png': None, '1811.02553v4-Figure1-1.png': None, '1811.02553v4-Figure10-1.png': None, '1811.02553v4-Figure11-1.png': None, '1811.02553v4-Figure12-1.png': None, '1811.02553v4-Figure13-1.png': None, '1811.02553v4-Figure14-1.png': None, '1811.02553v4-Figure15-1.png': None, '1811.02553v4-Figure16-1.png': None, '1811.02553v4-Figure17-1.png': None, '1811.02553v4-Figure18-1.png': None, '1811.02553v4-Figure19-1.png': None, '1811.02553v4-Figure2-1.png': None, '1811.02553v4-Figure20-1.png': None, '1811.02553v4-Figure3-1.png': None, '1811.02553v4-Figure4-1.png': None, '1811.02553v4-Figure5-1.png': None, '1811.02553v4-Figure6-1.png': None, '1811.02553v4-Figure7-1.png': None, '1811.02553v4-Figure8-1.png': None, '1811.02553v4-Figure9-1.png': None, '1811.02553v4-Table1-1.png': None, '1811.02721v3-Figure1-1.png': None, '1811.02721v3-Figure10-1.png': None, '1811.02721v3-Figure11-1.png': None, '1811.02721v3-Figure12-1.png': None, '1811.02721v3-Figure13-1.png': None, '1811.02721v3-Figure3-1.png': None, '1811.02721v3-Figure5-1.png': None, '1811.02721v3-Figure6-1.png': None, '1811.02721v3-Figure7-1.png': None, '1811.02721v3-Figure8-1.png': None, '1811.02721v3-Figure9-1.png': None, '1811.02721v3-Table1-1.png': None, '1811.02721v3-Table2-1.png': None, '1811.02721v3-Table3-1.png': None, '1811.02721v3-Table4-1.png': None, '1811.02721v3-Table5-1.png': None, '1811.02721v3-Table6-1.png': None, '1811.02721v3-Table7-1.png': None, '1811.02721v3-Table8-1.png': None, '1811.02721v3-Table9-1.png': None, '1811.06635v1-Figure1-1.png': None, '1811.06635v1-Figure2-1.png': None, '1811.06635v1-Figure3-1.png': None, '1811.06635v1-Table1-1.png': None, '1811.07073v3-Figure1-1.png': None, '1811.07073v3-Figure2-1.png': None, '1811.07073v3-Figure3-1.png': None, '1811.07073v3-Figure4-1.png': None, '1811.07073v3-Figure5-1.png': None, '1811.07073v3-Table1-1.png': None, '1811.07073v3-Table2-1.png': None, '1811.07073v3-Table3-1.png': None, '1811.07073v3-Table4-1.png': None, '1811.08257v1-Figure1-1.png': None, '1811.08257v1-Figure2-1.png': None, '1811.08257v1-Figure3-1.png': None, '1811.08257v1-Figure4-1.png': None, '1811.08257v1-Figure5-1.png': None, '1811.08257v1-Table2-1.png': None, '1811.08257v1-Table3-1.png': None, '1811.08257v1-Table4-1.png': None, '1811.08257v1-Table5-1.png': None, '1811.08257v1-Table6-1.png': None, '1811.08481v2-Figure1-1.png': None, '1811.08481v2-Figure10-1.png': None, '1811.08481v2-Figure2-1.png': None, '1811.08481v2-Figure3-1.png': None, '1811.08481v2-Figure4-1.png': None, '1811.08481v2-Figure5-1.png': None, '1811.08481v2-Figure6-1.png': None, '1811.08481v2-Figure8-1.png': None, '1811.08481v2-Figure9-1.png': None, '1811.08481v2-Table1-1.png': None, '1811.08481v2-Table2-1.png': None, '1811.08481v2-Table3-1.png': None, '1811.08481v2-Table4-1.png': None, '1811.08481v2-Table5-1.png': None, '1811.08481v2-Table7-1.png': None, '1811.08481v2-Table8-1.png': None, '1811.09393v4-Figure10-1.png': None, '1811.09393v4-Figure11-1.png': None, '1811.09393v4-Figure12-1.png': None, '1811.09393v4-Figure13-1.png': None, '1811.09393v4-Figure14-1.png': None, '1811.09393v4-Figure15-1.png': None, '1811.09393v4-Figure16-1.png': None, '1811.09393v4-Figure18-1.png': None, '1811.09393v4-Figure19-1.png': None, '1811.09393v4-Figure2-1.png': None, '1811.09393v4-Figure20-1.png': None, '1811.09393v4-Figure22-1.png': None, '1811.09393v4-Figure23-1.png': None, '1811.09393v4-Figure3-1.png': None, '1811.09393v4-Figure4-1.png': None, '1811.09393v4-Figure5-1.png': None, '1811.09393v4-Figure8-1.png': None, '1811.09393v4-Figure9-1.png': None, '1811.09393v4-Table2-1.png': None, '1811.09393v4-Table3-1.png': None, '1811.09393v4-Table4-1.png': None, '1811.09393v4-Table5-1.png': None, '1811.09393v4-Table6-1.png': None, '1811.10673v1-Figure10-1.png': None, '1811.10673v1-Figure2-1.png': None, '1811.10673v1-Figure3-1.png': None, '1811.10673v1-Figure4-1.png': None, '1811.10673v1-Figure5-1.png': None, '1811.10673v1-Figure6-1.png': None, '1811.10673v1-Figure7-1.png': None, '1811.10673v1-Figure8-1.png': None, '1811.10673v1-Figure9-1.png': None, '1811.10673v1-Table1-1.png': None, '1811.10673v1-Table2-1.png': None, '1812.00108v4-Figure1-1.png': None, '1812.00108v4-Figure2-1.png': None, '1812.00108v4-Table1-1.png': None, '1812.00108v4-Table2-1.png': None, '1812.00108v4-Table3-1.png': None, '1812.00281v3-Figure10-1.png': None, '1812.00281v3-Figure11-1.png': None, '1812.00281v3-Figure12-1.png': None, '1812.00281v3-Figure15-1.png': None, '1812.00281v3-Figure16-1.png': None, '1812.00281v3-Figure2-1.png': None, '1812.00281v3-Figure3-1.png': None, '1812.00281v3-Figure4-1.png': None, '1812.00281v3-Figure5-1.png': None, '1812.00281v3-Figure6-1.png': None, '1812.00281v3-Figure8-1.png': None, '1812.00281v3-Figure9-1.png': None, '1812.00281v3-Table1-1.png': None, '1812.00281v3-Table2-1.png': None, '1812.00281v3-Table3-1.png': None, '1812.00281v3-Table4-1.png': None, '1812.00281v3-Table6-1.png': None, '1812.00281v3-Table7-1.png': None, '1812.00281v3-Table8-1.png': None, '1812.06589v2-Figure1-1.png': None, '1812.06589v2-Figure2-1.png': None, '1812.06589v2-Figure3-1.png': None, '1812.06589v2-Figure4-1.png': None, '1812.06589v2-Figure5-1.png': None, '1812.06589v2-Figure6-1.png': None, '1812.06589v2-Table1-1.png': None, '1812.06589v2-Table2-1.png': None, '1812.06589v2-Table3-1.png': None, '1812.06589v2-Table4-1.png': None, '1812.10735v2-Figure1-1.png': None, '1812.10735v2-Figure2-1.png': None, '1812.10735v2-Figure3-1.png': None, '1812.10735v2-Figure4-1.png': None, '1812.10735v2-Figure5-1.png': None, '1812.10735v2-Figure6-1.png': None, '1812.10735v2-Table1-1.png': None, '1812.10735v2-Table2-1.png': None, '1812.10735v2-Table3-1.png': None, '1812.10735v2-Table4-1.png': None, '1901.00056v2-Figure1-1.png': None, '1901.00056v2-Figure2-1.png': None, '1901.00056v2-Figure3-1.png': None, '1901.00056v2-Table1-1.png': None, '1901.00056v2-Table2-1.png': None, '1901.00056v2-Table3-1.png': None, '1901.00056v2-Table4-1.png': None, '1901.00056v2-Table5-1.png': None, '1901.00056v2-Table6-1.png': None, '1901.00398v2-Figure1-1.png': None, '1901.00398v2-Figure10-1.png': None, '1901.00398v2-Figure11-1.png': None, '1901.00398v2-Figure12-1.png': None, '1901.00398v2-Figure2-1.png': None, '1901.00398v2-Figure3-1.png': None, '1901.00398v2-Figure4-1.png': None, '1901.00398v2-Figure5-1.png': None, '1901.00398v2-Figure6-1.png': None, '1901.00398v2-Figure7-1.png': None, '1901.00398v2-Figure8-1.png': None, '1901.00398v2-Figure9-1.png': None, '1901.00398v2-Table1-1.png': None, '1901.00398v2-Table2-1.png': None, '1901.00398v2-Table3-1.png': None, '1901.00398v2-Table4-1.png': None, '1901.00398v2-Table5-1.png': None, '1901.00398v2-Table6-1.png': None, '1901.00398v2-Table7-1.png': None, '1901.00398v2-Table8-1.png': None, '1901.00398v2-Table9-1.png': None, '1906.06589v3-Figure1-1.png': None, '1906.06589v3-Figure2-1.png': None, '1906.06589v3-Figure3-1.png': None, '1906.06589v3-Figure4-1.png': None, '1906.06589v3-Figure5-1.png': None, '1906.06589v3-Figure6-1.png': None, '1906.06589v3-Figure7-1.png': None, '1906.06589v3-Table1-1.png': None, '1906.06589v3-Table10-1.png': None, '1906.06589v3-Table11-1.png': None, '1906.06589v3-Table12-1.png': None, '1906.06589v3-Table2-1.png': None, '1906.06589v3-Table3-1.png': None, '1906.06589v3-Table4-1.png': None, '1906.06589v3-Table5-1.png': None, '1906.06589v3-Table6-1.png': None, '1906.06589v3-Table7-1.png': None, '1906.06589v3-Table8-1.png': None, '1906.06589v3-Table9-1.png': None, '1906.10843v1-Figure1-1.png': None, '1906.10843v1-Figure2-1.png': None, '1906.10843v1-Figure3-1.png': None, '1906.10843v1-Figure4-1.png': None, '1906.10843v1-Figure5-1.png': None, '1906.10843v1-Figure6-1.png': None, '1906.10843v1-Figure7-1.png': None, '1906.10843v1-Figure8-1.png': None, '1906.10843v1-Table1-1.png': None, '1906.10843v1-Table2-1.png': None, '1906.10843v1-Table3-1.png': None, '1906.10843v1-Table4-1.png': None, '1906.10843v1-Table5-1.png': None}, 'question': 'What are the main differences between the educational philosophies of the Bonaparte and Voltaire schools?'}


# %%

from datasets import load_dataset
dataset = load_dataset("PaulLerner/viquae_dataset",split='test',  trust_remote_code=True , streaming=True)
print(next(iter(dataset)))
# they some sort of repurposing here ->

# {'image': '512px-Jackie_Wilson.png',
#  'input': "this singer's re-issued song became the UK Christmas number one after helping to advertise what brand?",
#  'kilt_id': 'bb_8433',
#  'id': '02653c094f5ac04556a2ac853f7ce805',
#  'meta': {'left_context': '', 'mention': '', 'obj_surface': {'text': []}, 'partial_evidence': {'end_paragraph_id': [], 'meta': [], 'section': [], 'start_paragraph_id': [], 'title': [], 'wikipedia_id': []}, 'right_context': '', 'sub_surface': {'text': []}, 'subj_aliases': {'text': []}, 'template_questions': {'text': []}},
#  'original_question': "Jackie Wilson's re-issued song Reet Petite became the 1986 UK Christmas number one after helping to advertise what brand?",
#  'output': {'answer': ['Levi’s', 'Levi (brand', 'Levi Strauss Co.', 'Levi Strauss And Co.', "Levi's jeans", 'Go Forth Campaign', 'Levi (company)', 'Levi Strauss & Co.', 'Levi Strauss & Co. (LS&CO)', 'Fuerza Unida', 'Levi Strauss & Co', 'LS&Co', 'Levi Strauss And Company', 'Levi (jeans)', "Levi's", 'Levi-Strauss And Co', 'Levi 501', 'Levi Strauss & Company', 'Levi-Strauss & Co.', 'Levi Strauss And Co', 'Levi-Strauss And Co.', 'LS&Co.', 'levi strauss co', 'go forth campaign', 'levi strauss and company', 'ls co', 'levi s', 'levi jeans', 'levi strauss and co', 'fuerza unida', 'levi s jeans', 'levi company', 'levi strauss co ls co', 'levi strauss company', 'levi 501', 'levi brand'], 'meta': [], 'original_answer': "Levi's", 'provenance': [{'bleu_score': [1.0], 'end_character': [742], 'end_paragraph_id': [9], 'meta': [], 'section': ['Section::::Biography.:Early years and career.\n'], 'start_character': [736], 'start_paragraph_id': [9], 'title': ['Jackie Wilson'], 'wikipedia_id': ['149144']}]}, 'url': 'http://upload.wikimedia.org/wikipedia/commons/thumb/a/ae/Jackie_Wilson.png/512px-Jackie_Wilson.png', 'wikidata_id': 'Q181483'}

#%%

from datasets import load_dataset

dataset = load_dataset("merve/vqav2-small",split='validation',  trust_remote_code=True , streaming=True)
print(next(iter(dataset)))


#%%

from openpyxl.drawing.image import Image as openpyxl_image
from openpyxl import Workbook

import pandas as pd
from PIL import Image
import numpy as np
from io import BytesIO
from openpyxl.drawing.image import Image as openpyxl_image
from openpyxl import Workbook

# Load the dataset
dataset = load_dataset("HuggingFaceM4/AdVQA",split='validation',  trust_remote_code=True , streaming=True)

# Initialize an empty list for samples
samples = []

# Create an iterator from the dataset
dataset_iterator = iter(dataset)

# Get 10 unique samples
for _ in range(10):
    samples.append(next(dataset_iterator))

# Create an Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "QA Data"

# Loop through the samples
for i, sample in enumerate(samples):
    # Get the image, question, and answer
    image = sample['image']  # Assuming this is already a PIL Image object
    question = sample['question']
    answer = sample['answers'][0]['answer']

    # Resize the image to 250x250 pixels
    image = image.resize((250, 250))

    # Save the image temporarily in memory as bytes
    img_buffer = BytesIO()
    image.save(img_buffer, format='PNG')
    img_buffer.seek(0)

    # Insert question and answer into the worksheet
    ws.cell(row=i + 1, column=1, value=question)
    ws.cell(row=i + 1, column=2, value=answer)

    # Insert the image into the worksheet
    img_to_insert = openpyxl_image(img_buffer)
    img_to_insert.anchor = f'C{i + 1}'  # Position the image in the appropriate cell
    ws.add_image(img_to_insert)

# Save the Excel workbook
wb.save('answers.xlsx')

# %%
