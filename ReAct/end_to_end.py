#%%
import os
import openai
from openai import OpenAI
openai.api_key = os.environ["OPENAI_API_KEY"]
client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
import requests
import wikienv
import wrappers
import json
import sys
from PIL import Image
import io
import base64
import matplotlib.pyplot as plt
import random
import time
from prompts.instructions import *

#%%

# Define LLM functions
def llm(prompt, stop=["\n"]):
    response = client.chat.completions.create(
      model="gpt-4o-mini",
      messages=[
        {
            "role": "user",
            "content": prompt,
        }
      ],
      temperature=0,
      max_tokens=1024,
      top_p=1,
      frequency_penalty=0.0,
      presence_penalty=0.0,
      stop=stop
    )
    return response.choices[0].message.content

#%%
# Define environment and wrappers
env = wikienv.WikiEnv()
env = wrappers.AOKVQAWrapper(env, split="repurposed")
# env = wrappers.HotPotQAWrapperImage(env, split="repurposed")
# env = wrappers.HotPotQAWrapper(env, split="dev")
# env = wrappers.LoggingWrapper(env)

def step(env, action, image,last_thought):
    attempts = 0
    while attempts < 10:
        try:
            return env.step(action, image, last_thought)
        except requests.exceptions.Timeout:
            attempts += 1

#%%
# Load prompts
folder = './prompts/'
prompt_file = 'prompts_naive.json'
with open(folder + prompt_file, 'r') as f:
    prompt_dict = json.load(f)

webthink_examples = prompt_dict['aokvqa_human_annotated']
webthink_prompt = single_answer_image_cleaned + webthink_examples

#%%
# Define webthink function
def webthink(idx=None, prompt=webthink_prompt, to_print=True, image_dataset=True):
    (question, info) = env.reset(idx=idx, return_info=True)
    image = info['image']
    if to_print:
        print(idx, question)
    prompt += question + "\n"
    traj = ""
    n_calls, n_badcalls = 0, 0
    logs = []
    for i in range(1, 8):
        n_calls += 1
        if i==1 and image_dataset==True:
            image_pic = Image.open(io.BytesIO(base64.b64decode(image)))
            plt.imshow(image_pic)
            plt.show()
        thought_action = llm(prompt + f"Thought {i}:", stop=[f"\nObservation {i}:"])
        try:
            thought, action = thought_action.strip().split(f"\nAction {i}: ")
        except:
            n_badcalls += 1
            n_calls += 1
            thought = thought_action.strip().split('\n')[0]
            action = llm(prompt + f"Thought {i}: {thought}\nAction {i}:", stop=[f"\n"]).strip()
        obs, r, done, info = step(env, action = str(action[0].lower() + action[1:]), image=image, last_thought=thought)
        obs = obs.replace('\\n', '')
        step_str = f"Thought {i}: {thought}\nAction {i}: {action}\nObservation {i}: {obs}\n"
        prompt += step_str
        traj+=step_str
        if to_print:
            print(step_str)
        if done:
            break
    if not done:
        obs, r, done, info = step(env, "finish[]",  image=image, last_thought=thought)

    info.update({'n_calls': n_calls, 'n_badcalls': n_badcalls, 'traj': prompt})
    logs.append({
        'question': question,
        'answer': info['answer'],
        'gt_answer':info['gt_answer'],
        'reward': r,
        'n_calls': n_calls,
        'n_badcalls': n_badcalls,
        'almost_em': info['almost_em'],
        'f1': info['f1'],
        'em':info['em'],
        'trajectory':traj,
        'direct_answers': info['direct_answers'] if info['direct_answers'] else 0,
        'image': image,
    })
    return r, info, logs

#%%
# Run webthink on multiple indices
idxs = list(range(93))
random.Random(124).shuffle(idxs)

rs = []
infos = []
old_time = time.time()
logs=[]
for i in idxs:
    r, info, new_logs = webthink(i, to_print=True, image_dataset=True) # comment out image_dataset=True if you have to remove
    rs.append(info['em'])
    infos.append(info)
    logs.extend(new_logs) 
    print(info['almost_em'])
    print(sum(rs), len(rs), sum(rs) / len(rs), (time.time() - old_time) / len(rs))
    print('-----------')
    print()
# %%

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# Create a DataFrame from the logs
df = pd.DataFrame(logs)

# Create an Excel file
wb = Workbook()
ws = wb.active

# Write the header
for col, header in enumerate(df.columns):
    ws.cell(row=1, column=col + 1, value=header)

# Write the data
for row, (_, row_data) in enumerate(df.iterrows()):
    for col, value in enumerate(row_data):
        if col == 11:  # Assuming column index 9 is the image column
            # Decode the base64 string and save it as an image file
            image_data = base64.b64decode(value)
            image_stream = BytesIO(image_data)
            img = Image(image_stream)

            # Set the image size to 500x500 pixels
            img.width = 250
            img.height = 250

            # Insert the image into the cell
            ws.add_image(img, ws.cell(row=row + 2, column=col + 1).coordinate)
        else:
            ws.cell(row=row + 2, column=col + 1, value=value)


# Save the Excel file
wb.save("aokvqa_results_with_aokvqa_prompt.xlsx")


#%%