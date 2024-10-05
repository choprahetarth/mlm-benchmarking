#%%
import base64
import io
from datasets import load_dataset

import json
import numpy as np
import base64

dataset = load_dataset("HuggingFaceM4/A-OKVQA", split='train', trust_remote_code=True, streaming=True)
#%%

def transform_to_hotpotqa(example):
    """
    The format of output should be like this
    
    - id: a unique identifier for the question
    - question: the question being asked
    - answer: the answer to the question
    - type: the type of question (e.g. "bridge", "comparison", etc.) (still have not figured this out)
    - level: the difficulty level of the question
    - supporting_facts: a list of sentences that support the answer
    - context: a list of paragraphs related to the question
    
    The format of input should be like this 
    
    - image: image data
    - question_id: unique identifier for each question
    - question: text of the question being asked
    - choices: list of possible answers
    - correct_choice_idx: index of the correct answer in the choices list
    - direct_answers: list of direct answers to the question
    - difficult_direct_answer: boolean indicating whether the direct answer is difficult to determine
    - rationales: list of rationales or explanations for the answer
    """
    # Extract relevant information from the A-OKVQA example
    question_id = example['question_id']
    question = example['question']
    choices = example['choices']
    answer = example['choices'][example['correct_choice_idx']]
    rationales = example['rationales']
    direct_answers = example['direct_answers']

    # Convert PIL Image to base64 format
    buffered = io.BytesIO()
    example['image'].save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
    
    # Replace the image with its base64 string representation
    context = img_str 

    hotpotqa_example = {
        'question': question,
        'answer': answer,
        'direct_answers':direct_answers,
        'context': context  # Now the context is the base64-encoded image
    }

    return hotpotqa_example
#%%
# # Apply the transformation to the dataset and remove the 'image' column
hotpotqa_dataset = dataset.map(transform_to_hotpotqa, remove_columns=['image','question_id', 'rationales', 'difficult_direct_answer', 'correct_choice_idx', 'choices'])
#%%

# # Print the first example of the transformed dataset
print(json.dumps(next(iter(hotpotqa_dataset)), indent=1))

# %%
# Take 100 samples from the transformed dataset
samples = list(hotpotqa_dataset.take(100))

# Save the samples as a JSON file
with open('./data/aokvqa_repurposed_train.json', 'w') as f:
    json.dump(samples, f, indent=1)

# %%
# This code is for going through all samples of the validation dataset for annotating the few-shot samples
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import random
from datasets import load_dataset


# Load the dataset with the validation dataset for annotation
dataset = load_dataset("HuggingFaceM4/A-OKVQA", split='validation', trust_remote_code=True, streaming=True)
dataset = dataset.shuffle(seed=12)  # Shuffle the dataset
val_dataset = dataset.map(transform_to_hotpotqa, remove_columns=['image','question_id', 'rationales', 'difficult_direct_answer', 'correct_choice_idx', 'choices'])
val_samples = list(val_dataset.take(25))

# Create a DataFrame from the samples
df = pd.DataFrame(val_samples)

# Create an Excel file
wb = Workbook()
ws = wb.active

# Write the header
for col, header in enumerate(df.columns):
    ws.cell(row=1, column=col + 1, value=header)

# Write the data
for row, (_, row_data) in enumerate(df.iterrows()):
    for col, value in enumerate(row_data):
        if col == 3:  # Assuming column index 2 is the image column
            # Decode the base64 string and save it as an image files
            image_data = base64.b64decode(value)
            image_stream = BytesIO(image_data)
            img = Image(image_stream)

            # Set the image size to 500x500 pixels
            img.width = 250
            img.height = 250

            # Insert the image into the cell
            ws.add_image(img, ws.cell(row=row + 2, column=col + 1).coordinate)
        else:
            ws.cell(row=row + 2, column=col + 1, value=value)

# Save the Excel file
wb.save("output_val.xlsx")

# %%
